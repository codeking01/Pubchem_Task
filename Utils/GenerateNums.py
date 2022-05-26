"""
    -- coding: utf-8 --
    @Author: codeking
    @Data : 2022/5/3 12:20
    @File : GenerateNums.py
"""
# 这个用来生成序列数字，然后将其存入txt文件中
import openpyxl
from tqdm import tqdm

# 将读取的excel文件提取里面的cas生成txt文件
def GenTxt(excel_path):
    try:
        # 读取excel的某列，全部读取出来，然后根据正则去取出cas号
        read_path = excel_path
        # 读取的excel的对象
        wb = openpyxl.load_workbook(read_path)
    except Exception as e:
        print('未知错误,错误原因：', e)
    else:
        # 取第一张表名
        sheetnames = wb.sheetnames
        # 获取第一张表
        ws = wb[sheetnames[0]]
        rows = ws.max_row
        max_column = ws.max_column
        NumsList=[]
        # 依次取出里面的cas号
        for index in range(2, rows + 1):
            TheCid=int(ws.cell(row=index, column=2).value)
            NumsList.append(TheCid)
        IteratorLength=len(NumsList)
        # 路径和excel路径同
        txtpath=read_path.split('ALL_Excel')[0]
        txtname=read_path.split('/')[-1].split('.')[0]
        open(f'{txtpath}/ALL_Nums/{txtname}.txt', 'w').close()
        #将这个  NumsList内容写入txt
        for i in tqdm(NumsList):
            try:
                with open(f'{txtpath}/ALL_Nums/{txtname}.txt', encoding='utf-8', mode='a') as file:
                    file.write(str(i) + '\n')
                file.close()
            except Exception as e:
                print('文件生成失败，原因为：', e)
        print('结束写入')

def GenNums(txtpath,NumsList):
    # 把1到1000 写入txt文件
    # 根据传递的列表的值命名这个txt文件
    NumsList=NumsList.split(',')
    num_start=int(NumsList[0])
    num_end=int(NumsList[1])
    print('开始写入')
    # 先删除原来的文件再转化新的文件
    open(f'{txtpath}/nums.txt', 'w').close()
    for i in tqdm(range(num_start, num_end)):
        try:
            with open(f'{txtpath}/nums.txt', encoding='utf-8', mode='a') as file:
                file.write(str(i) + '\n')
            file.close()
        except Exception as e:
            print('文件生成失败，原因为：', e)
    print('结束写入')

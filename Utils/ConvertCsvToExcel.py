"""
    -- coding: utf-8 --
    @Author: codeking
    @Data : 2022/5/1 18:12
    @File : ConvertCsvToExcel.py
"""
#  由于csv文件操作存在一些问题，这里我转化为xlsx进行操作
import re

import openpyxl
from pandas import read_csv


def ConvertToExcel(path_name):
    csv_path = '{name}'.format(name=str(path_name))
    f = open(csv_path, encoding='utf-8')
    data = read_csv(f)
    path_name = path_name.split('.')[0]
    excel_path = '{path_name}.xlsx'.format(path_name=str(path_name))
    data.to_excel(excel_path)

# 存储cas号到excel中
def AddCasToExcel(excel_path):
    try:
        # 读取excel的某列，全部读取出来，然后根据正则去取出cas号
        read_path = excel_path
        wb = openpyxl.load_workbook(read_path)
    except Exception as e:
        print('路径可能有问题，请检查,错误原因：', e)
    else:
        # 取第一张表名
        sheetnames = wb.sheetnames
        # 获取第一张表
        ws = wb[sheetnames[0]]
        rows = ws.max_row
        max_column = ws.max_column
        # 获取所有cas的数据 用cas_list存储（列表）
        cas_list = []


        # 定义正则规则
        cas_pattern = re.compile(r"\|(\d{2,7}-\d{2}-\d{1})\||\|CAS-(\d{2,7}-\d{2}-\d{1})\|")
        # 存储cas到excel里面 cas数据在第4列
        for index in range(2, rows + 1):
            pre_cas = ws.cell(row=index, column=4).value

            # 部分的数据在excel里面没有
            if (pre_cas == None):
                cas_list.append('')
                ws.cell(index, max_column + 1, '')
            else:
                cas = cas_pattern.findall(pre_cas)
                # 考虑没有cas号的情况 存一个 '' 占位，确保顺序一致性，后面加判断
                if (len(cas) == 0):
                    cas_list.append('')
                    ws.cell(index, max_column + 1, '')
                else:
                    # 处理匹配到的数据  这个是这个类型  cas=[('71-55-6', ''), ('', '71-55-6')]
                    cas_temp_list = []
                    for cas_value_index in range(len(cas)):
                        cas_temp = cas[cas_value_index]
                        for c_value in cas_temp:
                            if (c_value != ''):
                                cas_temp_list.append(c_value)
                    # 删除重复的cas数据
                    cas = list(set(cas_temp_list))
                    # 在有cas的情况下，分别考虑多个cas和单个cas的情况
                    # 将'-'转换为'_'
                    cas = [i.replace('-', '_') for i in cas]
                    # 如果取到多个cas号 则以列表的形式存储  把列表的所有数据存储进去
                    if (len(cas) > 1):
                        temp_list = []
                        for storage_cas in cas:
                            # # 判断0开头的cas号，不是0才存进去
                            # if (storage_cas[0] != '0'):
                            temp_list.append(storage_cas)
                        cas_list.append(temp_list)
                        ws.cell(index, max_column + 1, '{data}'.format(data=temp_list))
                    else:
                        # cas列表可能为空
                        if (len(cas) == 0):
                            cas_list.append('')
                            ws.cell(index, max_column + 1, '')
                        else:
                            cas_list.append(cas[0])
                            ws.cell(index, max_column + 1, '{data}'.format(data=cas[0]))
        wb.save(excel_path)
        print("excel的cas全部存储完毕！")

# 存储cas号到excel中
def GetCaslist(excel_path):
    try:
        # 读取excel的某列，全部读取出来，然后根据正则去取出cas号
        read_path = excel_path
        wb = openpyxl.load_workbook(read_path)
    except Exception as e:
        print('路径可能有问题，请检查,错误原因：', e)
    else:
        # 取第一张表名
        sheetnames = wb.sheetnames
        # 获取第一张表
        ws = wb[sheetnames[0]]
        rows = ws.max_row
        max_column = ws.max_column
        # 获取所有cas的数据 用cas_list存储（列表）
        cas_list = []
        # 用来存储cid数据
        cid_list = []

        # 定义正则规则
        cas_pattern = re.compile(r"\|(\d{2,7}-\d{2}-\d{1})\||\|CAS-(\d{2,7}-\d{2}-\d{1})\|")
        # 存储cas到excel里面 cas数据在第4列
        for index in range(2, rows + 1):
            pre_cas = ws.cell(row=index, column=4).value
            # 将cid号全部取出来 存在一个列表里 cid数据在第2列
            cid_list.append(ws.cell(row=index, column=2).value)

            # 部分的数据在excel里面没有
            if (pre_cas == None):
                cas_list.append('')
                ws.cell(index, max_column + 1, '')
            else:
                cas = cas_pattern.findall(pre_cas)
                # 考虑没有cas号的情况 存一个 '' 占位，确保顺序一致性，后面加判断
                if (len(cas) == 0):
                    cas_list.append('')
                    ws.cell(index, max_column + 1, '')
                else:
                    # 处理匹配到的数据  这个是这个类型  cas=[('71-55-6', ''), ('', '71-55-6')]
                    cas_temp_list = []
                    for cas_value_index in range(len(cas)):
                        cas_temp = cas[cas_value_index]
                        for c_value in cas_temp:
                            if (c_value != ''):
                                cas_temp_list.append(c_value)
                    # 删除重复的cas数据
                    cas = list(set(cas_temp_list))
                    # 在有cas的情况下，分别考虑多个cas和单个cas的情况
                    # 将'-'转换为'_'
                    cas = [i.replace('-', '_') for i in cas]
                    # 如果取到多个cas号 则以列表的形式存储  把列表的所有数据存储进去
                    if (len(cas) > 1):
                        temp_list = []
                        for storage_cas in cas:
                            # # 判断0开头的cas号，不是0才存进去
                            # if (storage_cas[0] != '0'):
                            temp_list.append(storage_cas)
                        cas_list.append(temp_list)
                        ws.cell(index, max_column + 1, '{data}'.format(data=temp_list))
                    else:
                        # cas列表可能为空
                        if (len(cas) == 0):
                            cas_list.append('')
                            ws.cell(index, max_column + 1, '')
                        else:
                            cas_list.append(cas[0])
                            ws.cell(index, max_column + 1, '{data}'.format(data=cas[0]))
        return cas_list, cid_list